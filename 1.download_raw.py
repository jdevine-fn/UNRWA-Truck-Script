import gdown
import os
import platform
from datetime import datetime
import openpyxl

# ==================
# DOWNLOAD DATA FROM GOOGLE DRIVE WITHOUT MODIFICATION
# ==================

# Google Drive file ID for the input file (do not change)
file_id = '19oQZt7zWE29hK6Whnr9zop4gIGUValfxK14fQVHW18s'
download_url = f'https://drive.google.com/uc?id={file_id}'

# Get the current date in the format YYYYMMDD
current_date = datetime.now().strftime('%Y%m%d')

# Determine the user's desktop location (macOS and Windows compatibility)
if platform.system() == "Darwin":  # macOS
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
elif platform.system() == "Windows":  # Windows
    desktop = os.path.join(os.environ["HOMEPATH"], "Desktop")
else:
    raise Exception("Unsupported operating system. This script works on macOS and Windows only.")

# Create a folder on the desktop with the format "UNRWA Truck Data_YYYYMMDD"
folder_name = f"UNRWA Truck Data_{current_date}"
output_dir = os.path.join(desktop, folder_name)

# Ensure the directory exists
os.makedirs(output_dir, exist_ok=True)

# Define the final output file name
output_file = os.path.join(output_dir, "unrwa_trucks_raw.xlsx")

# Download the file using gdown (without any modifications to structure or content)
gdown.download(download_url, output_file, quiet=False, use_cookies=False, verify=False)

# The file is downloaded as-is, retaining the original structure, including multiple sheets and formatting.
print(f"File successfully downloaded and saved as {output_file} without any alterations.")

# Load the downloaded Excel file
wb = openpyxl.load_workbook(output_file)

# Check if 'Suppy Page' exists and rename it to 'Supply Page'
if 'Suppy Page' in wb.sheetnames:
    wb['Suppy Page'].title = 'Supply Page'
    print("Sheet 'Suppy Page' renamed to 'Supply Page'.")
else:
    print("No sheet named 'Suppy Page' found. No renaming done.")

# ==================
# PROCESS 'Quantity' COLUMN IN 'Supply Page' SHEET
# ==================

# Check if 'Supply Page' sheet exists
if 'Supply Page' in wb.sheetnames:
    ws = wb['Supply Page']
    print("Processing 'Supply Page' sheet to correct 'Quantity' column.")
else:
    print("Sheet 'Supply Page' not found. Exiting.")
    exit(1)

# Map header names to column letters
header_row = ws[1]  # Assuming headers are in the first row
header_map = {}  # Map from header name to column letter

for cell in header_row:
    header_map[cell.value] = cell.column_letter

# Check if 'Quantity' column exists
if 'Quantity' in header_map:
    quantity_col_letter = header_map['Quantity']
    quantity_col_index = openpyxl.utils.column_index_from_string(quantity_col_letter)
    print(f"'Quantity' column found at column {quantity_col_letter}.")
else:
    print("Column 'Quantity' not found in 'Supply Page'. Exiting.")
    exit(1)

# Process each cell in the 'Quantity' column starting from row 2
rows_processed = 0
date_cells_corrected = 0
errors_encountered = 0

for row in ws.iter_rows(min_row=2):
    cell = row[quantity_col_index - 1]  # zero-based index
    rows_processed += 1

    # Initialize quantity to None
    quantity = None

    try:
        if cell.value is None or cell.value == '':
            quantity = 0  # Preserve legitimate zeros or empty cells
        elif isinstance(cell.value, datetime):
            # Cell is a datetime object, likely due to date formatting error
            # Compute the date serial number
            base_date = datetime(1899, 12, 31)
            delta = cell.value - base_date
            quantity = delta.days
            date_cells_corrected += 1
        elif isinstance(cell.value, (int, float)):
            quantity = cell.value  # Cell contains a numeric value
        else:
            # Try to convert cell.value to float
            quantity = float(cell.value)
    except Exception as e:
        print(f"Error processing cell {cell.coordinate}: {e}")
        errors_encountered += 1
        quantity = 0  # Default to zero or handle as needed

    # Update the cell with the corrected quantity
    cell.value = quantity
    # Set the cell's number format to 'General' to avoid future misinterpretation
    cell.number_format = 'General'

# Save the changes to the file
wb.save(output_file)
print(f"Processing complete. {rows_processed} rows processed.")
print(f"{date_cells_corrected} cells in 'Quantity' column corrected from date format to numeric values.")
if errors_encountered > 0:
    print(f"{errors_encountered} errors encountered during processing.")
else:
    print("No errors encountered during processing.")
